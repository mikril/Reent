from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font
import csv
import re
from jinja2 import Environment, FileSystemLoader
from datetime import datetime
import matplotlib.pyplot as plt
import numpy as np
import pdfkit
from prettytable import PrettyTable
import os
import doctest


"""Глобальные словари

Globals:
    expiriences: перевод количиства опыта работы на русский язык
    currencies: перевод названия валюты на русский язык
    fieldToRus: перевод названий столбцов на русский
    filterToNames: обратный перевод fieldToRus
    currency_to_rub: перевод всех валют в рублиa
"""
expiriences={
"noExperience": "Нет опыта",
"between1And3": "От 1 года до 3 лет",
"between3And6": "От 3 до 6 лет",
"moreThan6": "Более 6 лет"
}

currencies={
"AZN": "Манаты",
"BYR": "Белорусские рубли",
"EUR": "Евро",
"GEL": "Грузинский лари",
"KGS": "Киргизский сом",
"KZT": "Тенге",
"RUR": "Рубли",
"UAH": "Гривны",
"USD": "Доллары",
"UZS": "Узбекский сум"}

fieldToRus = {
"name": "Название",
"description": "Описание",
"key_skills": "Навыки",
"experience_id": "Опыт работы",
"premium": "Премиум-вакансия",
"employer_name": "Компания",
"currency":"Оклад",
"area_name": "Название региона",
"published_at": "Дата публикации вакансии"
}

filterToNames = {
"Название": "name",
"Описание":"description",
"Компания":"employer_name",
"Оклад":"salary_from",
"Дата публикации вакансии":"published_at",
"Опыт работы": "experience_id",
"Навыки":"key_skills",
"Премиум-вакансия":"premium",
"Идентификатор валюты оклада":"salary_currency",
"Название региона":"area_name"
}

currency_to_rub = {  
    "Манаты": 35.68,  
    "Белорусские рубли": 23.91,  
    "Евро": 59.90,  
    "Грузинский лари": 21.74,  
    "Киргизский сом": 0.76,  
    "Тенге": 0.13,  
    "Рубли": 1,  
    "Гривны": 1.64,  
    "Доллары": 60.66,  
    "Узбекский сум": 0.0055,  
}

class InputConect:
    """ Класс InputConect для ввода значений, их проверки и отображения в консоли

        Attributes: 
            whatPrint(str): То что просит пользователь как вывод
        
    """
    def __init__(self,whatPrint):
        """ Иницилизирует InputConect, проверяет что нужно вводить в зависимости от того что нужно вывести

            Args:
                whatPrint(str): То что просит пользователь как вывод

            
        """
        if whatPrint=="Статистика":
            self.file = input("Введите название файла: ")
            self.filterElements=("Название: "+input("Введите название профессии: ")).split(": ")
            self.sortElements=""
            self.reversVacancies=  ""
            self.fromTo=""
            self.names=""
        if whatPrint=="Вакансии":
            self.file = input("Введите название файла: ")
            self.filterElements=input("Введите параметр фильтрации: ").split(": ")
            self.sortElements=input("Введите параметр сортировки: ")
            self.reversVacancies=  input("Обратный порядок сортировки (Да / Нет): ")
            self.fromTo=input("Введите диапазон вывода: ").split()
            self.names=input("Введите требуемые столбцы: ")
     

        
    def checkInput(self):
        """ Проверка введеных значений на корректность
        """
        filterElements=self.filterElements
        sortElements=self.sortElements
        if len(filterElements)==1 and len(filterElements[0])!=0:
            print("Формат ввода некорректен")
            exit()
        elif not(filterElements[0] in filterToNames ) and len(filterElements[0])!=0:
            print("Параметр поиска некорректен")  
            exit()  
        if len(sortElements)!=0 and not(sortElements in list(filterToNames.keys())):
            print("Параметр сортировки некорректен")
            exit()
        if self.reversVacancies=="Да":
            self.reversVacancies=True
        elif self.reversVacancies=="Нет":
            self.reversVacancies=False
        elif len(self.reversVacancies)==0:
            self.reversVacancies=False
        else:
            print("Порядок сортировки задан некорректно")
            exit()
   
    def print_vacancies(self, data_vacancies, dic_naming):
        """ Вывод итоговой таблицы со значениями в консоли 
            
            Args:
                data_vacancies(list): Cписок вакансий
                dic_naming(list): Список названий столбцов
        """
        if(len(self.fromTo)==2):
            end=int(self.fromTo[1])-1
            start=int(self.fromTo[0])-1
        if(len(self.fromTo)==1):
            start=int(self.fromTo[0])-1
            end=int(len(data_vacancies.vacancies_objects))
        if(len(self.fromTo)==0):
            start=0
            end=int(len(data_vacancies.vacancies_objects))
        
        x=PrettyTable()
        counter=0

        x.field_names=["№"]+list(dic_naming.values())
        x._max_width = {"Название":20, "Описание":20, "Навыки":20, "Опыт работы":20, "Премиум-вакансия":20, "Компания":20, "Оклад":20, "Название региона":20, "Дата публикации вакансии":20}
        x.hrules=1
        x.align="l"
        vacancy=start
        boolka=True

    
        while (vacancy < end):
            formatVacancy=data_vacancies.vacancies_objects[vacancy]
            boolka=False
            val=[str(vacancy+1)]
            for item in formatVacancy.elements:
                if type(item)==list:
                    element = '\n'.join(str(x) for x in item)
                elif type(item)==Salary:
                    element=item.salary
                elif  type(item)== datetime:
                    element="{:%d.%m.%Y}".format(item)
                else:
                    element=item
                val.append(element[0:100]+"..." if len(element)>100 else element)
            x.add_row(val)
            vacancy+=1
        if (boolka):            
            print("Ничего не найдено")
            exit()

        if(len(self.names)>0):
            print(x.get_string(fields=["№"]+self.names.split(", ")))
        else:
            print(x)
     
        

class DataSet:
    """ Класс DataSet для обработки, филтрации, сортировки вакансий

        Attributes: 
            file_name(str): Название файла
            filterElements(list): Два элемента столбец по которому фильтровать, значение филтрации
            sortElements(str): Столбец который нужно сортировать
            reversVacancies(str): Отреверсировать сортировку Да, Нет?
            vacancies_objects(list): контейнер для измененых, отфильтрованных, отсортированных, вакансий
    """
    def __init__(self,file_name,filterElements,sortElements,reversVacancies,vacancies_objects=[]):
        """ Иницилизирует DataSet 

        Args: 
            file_name(str): Название файла
            filterElements(list): Два элемента столбец по которому фильтровать, значение филтрации
            sortElements(str): Столбец который нужно сортировать
            reversVacancies(str): Отреверсировать сортировку Да, Нет?
            vacancies_objects(list): контейнер для измененых, отфильтрованных, отсортированных, вакансий

            >>> type(DataSet('vacancies.csv',["Дата публикации","15.12.2022"],"Думайте","")).__name__
            'DataSet'
            >>> DataSet('vacancies.csv',["Дата публикации","15.12.2022"],"Думайте","").file_name 
            'vacancies.csv'
            >>> DataSet('vacancies.csv',["Дата публикации","15.12.2022"],"Думайте","").filterElements[0] 
            'Дата публикации'
            >>> DataSet('vacancies.csv',["Дата публикации","15.12.2022"],"Думайте","").filterElements[1] 
            '15.12.2022'
            >>> DataSet('vacancies.csv',["Дата публикации","15.12.2022"],"Думайте","").sortElements 
            'Думайте'
            >>> DataSet('vacancies.csv',["Дата публикации","15.12.2022"],"Думайте","").reversVacancies 
            ''
        """
        self.file_name=file_name
        self.vacancies_objects=vacancies_objects
        self.filterElements=filterElements
        self.sortElements=sortElements
        self.reversVacancies=reversVacancies

    def correctVacanceis(self):
        """ Функция чтобы работала другая функция(без неё не работает)
        """
        return self.checkEmpty()
    
    def checkEmpty(self): 
        """ Функция запуска чтения файла, проверки файла на пустоту вцелом, на пустоту количества вакансий

            Returns: 
                DataSet: либо прекращает выполнение программы, либо возвращает заполненый DataSet 
        """
        reader, list_naming=self.сsv_reader()
        if len(list_naming)>0 and len(reader)>0:
            return self.csv_ﬁler(reader,list_naming)
        elif len(list_naming)>0:
            print("Нет данных")
            exit()
        else:
            print("Пустой файл")  
            exit()  
        
    def сsv_reader(self):
        """ Функция для чтения файла, разбитие его на массив вакансий, на масив названий столбцов

            Returns:
                list: список вакансий 
                list: список названий столбцов элементов вакансии
        """
        file_name=self.file_name
        list_naming=[]
        reader=[]
        with open(ﬁle_name, encoding="utf-8-sig") as File: 
            for row in csv.reader(File, delimiter=',', quoting=csv.QUOTE_MINIMAL):
                if(not("" in row)):
                    if (len(list_naming) > 0  and len(list_naming)<=len(row)):
                        reader.append(row)
                    if (len(list_naming)==0):
                        list_naming = row
        return reader, list_naming   
    def csv_ﬁler( self, reader, list_naming):  
        """ Функция перезаписи двух списков в список словарей, для запуска функции форматирования элементов вакансии, фильтрации и сортировки этого словарая

            Args:
                reader: список вакансий 
                list_naming: список названий столбцов элементов вакансии

            Returns:
                DataSet: возвращает заполненый DataSet 

            >>> DataSet('vacancies.csv',["Дата публикации вакансии","15.12.2022"],"","").csv_ﬁler([['папич','быть величайшим','дота казик','moreThan6','Да','ютюб','Винница','2022-05-31T17:32:31+0300'],['monkey', 'asdaisfuiasd', 'banana', 'noExperience', 'Да', 'zoo', 'Moscow', '2022-12-15T17:32:31+0300']],["name","description","key_skills","experience_id","premium","employer_name","area_name","published_at"]).vacancies_objects[0].elements
            ['monkey', 'asdaisfuiasd', 'banana', 'Нет опыта', 'Да', 'zoo', '', 'Moscow', datetime.datetime(2022, 12, 15, 17, 32, 31, 30000)]
        """
        for element in reader:
            dictVacancy={}
            for i in range(len(element)):
                dictVacancy[list_naming[i]]=element[i]
            self.vacancies_objects.append(self.formatter(dictVacancy))
        #фильтрация
        if len(self.filterElements[0])>0:
            if filterToNames[self.filterElements[0]] in list(filterToNames.values()): 

                if filterToNames[self.filterElements[0]]=="salary_from":
                    self.vacancies_objects=list(filter(lambda x:   float(x.salary.salary_from.replace(" ",""))<=float(self.filterElements[1]) and float(x.salary.salary_to.replace(" ",""))>=float(self.filterElements[1]),self.vacancies_objects))
                elif self.filterElements[0]=="Навыки": 
                    self.vacancies_objects=list(filter(lambda x:   set(self.filterElements[1].split(", ")).issubset(x.key_skills),self.vacancies_objects))
                elif filterToNames[self.filterElements[0]]=="salary_currency":
                    self.vacancies_objects=list(filter(lambda x:   self.filterElements[1]==x.salary.salary_currency,self.vacancies_objects))
                elif filterToNames[self.filterElements[0]]=="published_at":
                    self.vacancies_objects=list(filter(lambda x:   self.filterElements[1]=="{:%d.%m.%Y}".format(x.published_at),self.vacancies_objects))
                else: 
                    self.vacancies_objects=list(filter(lambda x:  self.filterElements[1] in getattr(x, filterToNames[self.filterElements[0]]), self.vacancies_objects))
                    #self.vacancies_objects=list(filter(lambda x:  getattr(x, filterToNames[self.filterElements[0]]) == self.filterElements[1], self.vacancies_objects))    
        #сортировка
        if self.sortElements=="Название" or self.sortElements=="Описание" or  self.sortElements=="Компания" or self.sortElements=="Название региона" or self.sortElements=="Дата публикации вакансии" or self.sortElements=="Премиум-вакансия":
            self.vacancies_objects.sort( key=lambda x: getattr(x,list(fieldToRus.keys())[list(fieldToRus.values()).index(self.sortElements)]), reverse=self.reversVacancies)
        if self.sortElements=="Оклад":
            self.vacancies_objects.sort( key=lambda x: (float(x.salary.salary_from.replace(" ",""))+float(x.salary.salary_to.replace(" ","")))*currency_to_rub[x.salary.salary_currency], reverse=self.reversVacancies)
        if self.sortElements=="Навыки":
            
            self.vacancies_objects.sort(key=lambda x: len(x.key_skills) if type(x.key_skills)==list else 1, reverse=self.reversVacancies)
            
        if self.sortElements=="Опыт работы":
            self.vacancies_objects.sort( key=lambda x: list(expiriences.values()).index(x.experience_id), reverse=self.reversVacancies)
        return self
    
    def yearDinamic(self,name):
        """ Функция создания динамики зарплат по годам, количество вакансий по годам, зарплат по городам, количества вакансий по городам, зарплат по годам для конкретной вакансии, количество вакансий по годам для конкретной вакансии

            Args:
                name: название вакансии
                

            Returns:
                dict: динамика зарплат по годам
                dict: динамика количество вакансий по годам
                dict: динамика зарплат по городам
                dict: динамика количества вакансий по городам
                dict: динамика зарплат по годам для конкретной вакансии
                dict: динамика количество вакансий по годам для конкретной вакансии

        """
        countVacancyesYear={}
        salarysYear={}
        salaryTown={}
        VacanciesTown={}
        filterSalarysYear={}
        filterCountVacancyesYear={}
        
        for vacancy in self.vacancies_objects: 
            date=int("{:%Y}".format(vacancy.published_at))
            if not(date in countVacancyesYear):
                countVacancyesYear[date]=1
                filterCountVacancyesYear[date]=0
            else:
                countVacancyesYear[date]+=1
            if not(date in salarysYear):
                filterSalarysYear[date]=0
                salarysYear[date]=(float(vacancy.salary.salary_to.replace(" ",""))+float(vacancy.salary.salary_from.replace(" ","")))/2*currency_to_rub[vacancy.salary.salary_currency]
            else:
                salarysYear[date]+=(float(vacancy.salary.salary_to.replace(" ",""))+float(vacancy.salary.salary_from.replace(" ","")))/2*currency_to_rub[vacancy.salary.salary_currency]
            if name in vacancy.name:
                if (date in filterCountVacancyesYear):
                    filterCountVacancyesYear[date]+=1
                if (date in filterSalarysYear):
                    filterSalarysYear[date]+=(float(vacancy.salary.salary_to.replace(" ",""))+float(vacancy.salary.salary_from.replace(" ","")))/2*currency_to_rub[vacancy.salary.salary_currency]
            
            if not(vacancy.area_name in VacanciesTown):
                VacanciesTown[vacancy.area_name]=1
            else:
                VacanciesTown[vacancy.area_name]+=1
            if not(vacancy.area_name in salaryTown):
                salaryTown[vacancy.area_name]=(float(vacancy.salary.salary_to.replace(" ",""))+float(vacancy.salary.salary_from.replace(" ","")))/2*currency_to_rub[vacancy.salary.salary_currency]
            else:
                salaryTown[vacancy.area_name]+=(float(vacancy.salary.salary_to.replace(" ",""))+float(vacancy.salary.salary_from.replace(" ","")))/2*currency_to_rub[vacancy.salary.salary_currency]
        salarysYear=dict(sorted(salarysYear.items(), key=lambda x: x[0]))
        countVacancyesYear=dict(sorted(countVacancyesYear.items(), key=lambda x: x[0]))
        filterSalarysYear=dict(sorted(filterSalarysYear.items(), key=lambda x: x[0]))
        filterCountVacancyesYear=dict(sorted(filterCountVacancyesYear.items(), key=lambda x: x[0]))
        VacanciesTown=dict(sorted(VacanciesTown.items(), key=lambda x: x[1], reverse=True))
        salaryTown=dict(sorted(salaryTown.items(), key=lambda x: x[1]/VacanciesTown[x[0]],reverse=True))

        salaryTown=dict(filter(lambda x: VacanciesTown[x[0]]/len(self.vacancies_objects) >=0.01, salaryTown.items()))
        return salarysYear,countVacancyesYear,VacanciesTown,salaryTown,filterSalarysYear,filterCountVacancyesYear
   
    def formatter(self,row): 
        """ Форматирует элемнты вакансии

            Args:
                row: словарь вакансии

            Returns:
                Vacancy: возвращает отформаттированную вакансию

            >>> DataSet('vacancies.csv',["Дата публикации","15.12.2022"],"Думайте","").formatter({"name":"monkey","description":"<p>asdaisfuiasd</p>","key_skills":"banana","experience_id":"noExperience","premium":"Да","employer_name":"zoo","area_name":"Moscow","published_at":"2022-05-31T17:32:31+0300"}).elements
            ['monkey', 'asdaisfuiasd', 'banana', 'Нет опыта', 'Да', 'zoo', '', 'Moscow', datetime.datetime(2022, 5, 31, 17, 32, 31, 30000)]
        """
        args=["","","","", "","","","",""]
        namesindex=["name","description","key_skills","experience_id","premium","employer_name","salary","area_name","published_at"]
        argsSalary=["","","",""]
        namessalary=["salary_from","salary_to","salary_gross","salary_currency"]
        keys=list(row.keys())
        if len(keys)>0:
            for i in range(len(keys)):
                value = row[keys[i]]
                
                if("\n" in value):
                    value = value.split("\n")
                    for index in range(len(value)):
                        value[index] = ' '.join(re.sub(r"<[^>]+>", '', value[index]).split())
                else:
                    value = ' '.join(re.sub(r"<[^>]+>", '', value).split())
                
                if(type(value) != list and value.upper()=="TRUE"):
                    value = "Да"
                if(type(value) != list and value.upper()=="FALSE"):
                    value = "Нет"
                if (keys[i]=="experience_id"):
                    value= expiriences[value]
                if (keys[i]=="published_at"):
                    a = datetime.strptime(value.replace("+",".").replace("T"," "), '%Y-%m-%d %H:%M:%S.%f')
                    value= a

                if(keys[i]=="salary_from"):
                    value='{:,}'.format(int(float(row[keys[i]]))).replace(',', ' ')
                    argsSalary[namessalary.index(keys[i])]=value
                elif(keys[i]=="salary_to"):
                    value='{:,}'.format(int(float(row[keys[i]]))).replace(',', ' ')
                    argsSalary[namessalary.index(keys[i])]=value
                elif(keys[i]=="salary_gross"):
                    value = (("С вычетом налогов") if (row[keys[i]].upper()=="FALSE") else ("Без вычета налогов"))
                    argsSalary[namessalary.index(keys[i])]=value
                elif(keys[i]=="salary_currency"):
                    value = currencies[row[keys[i]]]
                    argsSalary[namessalary.index(keys[i])]=value
                    args[namesindex.index("salary")]=Salary(*argsSalary)
                else:
                    args[namesindex.index(keys[i])]=value
        
        result=Vacancy(*args)
        return result
    

    
class Vacancy:
    """ Класс Vacancy для хранения элемнтов вакансии

        Attributes: 
            name(str): Название вакансии
            description(str): Описание
            key_skills(list): Список скиллов
            experience_id(str): Опыт работы
            premium(str): Премиум вауансия Да, Нет?
            employer_name(str):Название компании
            salary(str): Сюда должен передаваться класс зарплаты
            area_name(str): Название города
            published_at(str): Дата публикации вакансии
    """
    def __init__(self,name="",description="",key_skills="",experience_id="",premium="",employer_name="",salary="",area_name="",published_at=""):
        """ Инициализирует Vacancy, объединяет элементы в список

            Args:
                name(str): Название вакансии
                description(str): Описание
                key_skills(list): Список скиллов
                experience_id(str): Опыт работы
                premium(str): Премиум вауансия Да, Нет?
                employer_name(str):Название компании
                salary(str): Сюда должен передаваться класс оклада
                area_name(str): Название города
                published_at(str): Дата публикации вакансии
            >>> type(Vacancy()).__name__
            'Vacancy'
            >>> Vacancy('папич','быть величайшим',['дота','казик'],'всю жизнь','ДЫА','ютуб',Salary('1000000','infinity','Нет','RUB'),'Винница','07.07.777').name 
            'папич'
            >>> Vacancy('папич','быть величайшим',['дота','казик'],'всю жизнь','ДЫА','ютуб',Salary('1000000','infinity','Нет','RUB'),'Винница','07.07.777').description 
            'быть величайшим'
            >>> Vacancy('папич','быть величайшим',['дота','казик'],'всю жизнь','ДЫА','ютуб',Salary('1000000','infinity','Нет','RUB'),'Винница','07.07.777').key_skills 
            ['дота', 'казик']
            >>> Vacancy('папич','быть величайшим',['дота','казик'],'всю жизнь','ДЫА','ютуб',Salary('1000000','infinity','Нет','RUB'),'Винница','07.07.777').experience_id 
            'всю жизнь'
            >>> Vacancy('папич','быть величайшим',['дота','казик'],'всю жизнь','ДЫА','ютуб',Salary('1000000','infinity','Нет','RUB'),'Винница','07.07.777').premium 
            'ДЫА'
            >>> Vacancy('папич','быть величайшим',['дота','казик'],'всю жизнь','ДЫА','ютуб',Salary('1000000','infinity','Нет','RUB'),'Винница','07.07.777').employer_name
            'ютуб'
            >>> Vacancy('папич','быть величайшим',['дота','казик'],'всю жизнь','ДЫА','ютуб',Salary('1000000','infinity','Нет','RUB'),'Винница','07.07.777').area_name 
            'Винница'
            >>> Vacancy('папич','быть величайшим',['дота','казик'],'всю жизнь','ДЫА','ютуб',Salary('1000000','infinity','Нет','RUB'),'Винница','07.07.777').published_at 
            '07.07.777'
            >>> len(Vacancy('папич','быть величайшим',['дота','казик'],'всю жизнь','ДЫА','ютуб',Salary('1000000','infinity','Нет','RUB'),'Винница','07.07.777').elements)
            9
        """
        self.name=name
        self.description=description 
        self.key_skills=key_skills
        self.experience_id=experience_id
        self.premium=premium
        self.employer_name=employer_name
        self.salary=salary
        self.area_name=area_name
        self.published_at=published_at
        self.elements=[name,description,key_skills,experience_id,premium,employer_name,salary,area_name,published_at]
        

class Salary:
    """ Класс Salary для хранения элемнтов оклада

        Attributes: 
            salary_from(str): Нижняя граница вилки отклада
            salary_to(str): Верхняя граница вилки оклада
            salary_gross(list): С вычетом, без вычетов налога
            salary_currency(str): Валюта оклада
    """
    def __init__(self,salary_from ="",salary_to ="",salary_gross ="",salary_currency =""):
        """ Инициализирует Salary, создает обьединённую строку

            Args: 
                salary_from(str): Нижняя граница вилки отклада
                salary_to(str): Верхняя граница вилки оклада
                salary_gross(list): С вычетом, без вычетов налога
                salary_currency(str): Валюта оклада

            >>> type(Salary()).__name__
            'Salary'
            >>> Salary('1000000','infinity','Нет','RUB').salary_from 
            '1000000'
            >>> Salary('1000000','infinity','Нет','RUB').salary_to 
            'infinity'
            >>> Salary('1000000','infinity','Нет','RUB').salary_gross 
            'Нет'
            >>> Salary('1000000','infinity','Нет','RUB').salary_currency 
            'RUB'
        """
        self.salary_from=salary_from
        self.salary_to=salary_to
        self.salary_gross=salary_gross
        self.salary_currency=salary_currency
        self.salary=salary_from+" - "+salary_to+" (" +salary_currency +") (" +salary_gross+")"

class Report:
    """ Класс Report для создания диаграмм, файла эксель, файла pdf 

        Attributes: 
            name(str): Название вакансии
            salarysYear: динамика зарплат по годам
            countVacancyesYear: динамика количество вакансий по годам
            filterSalarysYear: динамика зарплат по городам
            filterCountVacancyesYear: динамика количества вакансий по городам
            salaryTown: динамика зарплат по годам для конкретной вакансии
            upgradeVacanciesTown: динамика количество вакансий по годам для конкретной вакансии
    """
    def __init__(self, name,salarysYear,countVacancyesYear,filterSalarysYear,filterCountVacancyesYear,salaryTown,upgradeVacanciesTown ):
        """ Инициализирует Report 

            Args: 
                name(str): Название вакансии
                salarysYear(list): динамика зарплат по годам
                countVacancyesYear(list): динамика количество вакансий по годам
                filterSalarysYear(list): динамика зарплат по городам
                filterCountVacancyesYear(list): динамика количества вакансий по городам
                salaryTown(list): динамика зарплат по годам для конкретной вакансии
                upgradeVacanciesTown(list): динамика количество вакансий по годам для конкретной вакансии
            >>> type(Report('программист',[{2022:1000000}],[{2022:1}],[{"Винница":1000000}],[{"Винница":1}],[{2022:1000000}],[{2022:1}])).__name__
            'Report'
            >>> Report('программист',[{2022:1000000}],[{2022:1}],[{"Винница":1000000}],[{"Винница":1}],[{2022:1000000}],[{2022:1}]).name
            'программист'
            >>> Report('программист',[{2022:1000000}],[{2022:1}],[{"Винница":1000000}],[{"Винница":1}],[{2022:1000000}],[{2022:1}]).salarysYear[0]
            {2022: 1000000}
            >>> Report('программист',[{2022:1000000}],[{2022:1}],[{"Винница":1000000}],[{"Винница":1}],[{2022:1000000}],[{2022:1}]).countVacancyesYear[0][2022]
            1
            >>> Report('программист',[{2022:1000000}],[{2022:1}],[{"Винница":1000000}],[{"Винница":1}],[{2022:1000000}],[{2022:1}]).filterSalarysYear[0]["Винница"]
            1000000
            >>> list(Report('программист',[{2022:1000000}],[{2022:1}],[{"Винница":1000000}],[{"Винница":1}],[{2022:1000000}],[{2022:1}]).filterCountVacancyesYear[0].keys())[0]
            'Винница'
            >>> Report('программист',[{2022:1000000}],[{2022:1}],[{"Винница":1000000}],[{"Винница":1}],[{2022:1000000}],[{2022:1}]).salaryTown
            [{2022: 1000000}]
            >>> Report('программист',[{2022:1000000}],[{2022:1}],[{"Винница":1000000}],[{"Винница":1}],[{2022:1000000}],[{2022:1}]).upgradeVacanciesTown
            [{2022: 1}]
        """
        self.name=name
        self.salarysYear=salarysYear
        self.countVacancyesYear=countVacancyesYear
        self.filterSalarysYear=filterSalarysYear
        self.filterCountVacancyesYear=filterCountVacancyesYear
        self.salaryTown=salaryTown
        self.upgradeVacanciesTown=upgradeVacanciesTown
   
    def generate_diagrams(self):
        """Создаёт диаграммы
        """
        plt.rcParams.update({'font.size': 8})
        
        labels = self.salarysYear.keys()
        sredn_slary = self.salarysYear.values()
        name_salary = self.filterSalarysYear.values()

        x = np.arange(len(labels))  # the label locations
        width = 0.35  # the width of the bars

        ax = plt.subplot(221)
        ax.bar(x - width/2, sredn_slary, width, label='средняя з/п')
        ax.bar(x + width/2, name_salary, width, label='з/п программист')

        # Add some text for labels, title and custom x-axis tick labels, etc.
        ax.set_title('Уровень зарплат по городам')
        ax.set_xticks(x, labels, rotation =90)
        ax.legend()

        sredn_slary = self.countVacancyesYear.values()
        name_salary = self.filterCountVacancyesYear.values()

        bx = plt.subplot(222)
        bx.bar(x - width/2, sredn_slary, width, label='Количество вакансий')
        bx.bar(x + width/2, name_salary, width, label='Количество вакансий программист')

        # Add some text for labels, title and custom x-axis tick labels, etc.
        bx.set_title('Количестов вакансий по годам')
        bx.set_xticks(x, labels,rotation =90)
        bx.legend()
        

        cx = plt.subplot(223)

        # Example data
        towns = salaryTown.keys()
        y_pos = np.arange(len(towns))
        salarys = salaryTown.values()
        error = np.random.rand(len(towns))

        cx.barh(y_pos, salarys, xerr=error, align='center')
        cx.set_yticks(y_pos, labels=towns)
        cx.invert_yaxis()  # labels read top-to-bottom
        cx.set_title('Уровень зарплат по городам')
       
        upgradeVacanciesTown["Другие"]=1-sum(upgradeVacanciesTown.values())
        labels = upgradeVacanciesTown.keys()
        sizes =upgradeVacanciesTown.values()
        plt.subplots_adjust(wspace=0.5, hspace=0.5)
        
        dx = plt.subplot(224)
        dx.pie(sizes, labels=labels)
        dx.set_title('Доля вакансий по городам')
        upgradeVacanciesTown.pop("Другие")
        plt.savefig('graph.png', dpi = 200)
    
    def do_border(self, ws, cell_range):
        """Создаёт гранницы для таблицы

            Args: 
                ws(worksheet): Активный лист в эксель
                cell_range(str): Активные ячейки
        """
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top = thin, left = thin, right = thin, bottom = thin)

    def do_bold(self, ws, cell_range):
        """Задаёт толщину тексту

            Args: 
                ws(worksheet): Активный лист в эксель
                cell_range(str): Активные ячейки
        """
        for row in ws[cell_range]:
            for cell in row:
                cell.font = Font(bold = True) 
    def do_width_sizes(self, ws):
        """Задаёт размеры всем столбцам
            Args: 
                ws(worksheet): Активный лист в эксель   
        """
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)))) + 0.5
        for col, value in dims.items():
            ws.column_dimensions[col].width = value   
    def generate_excel(self):
        """Создаёт эксель файл
        """
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Статистика по годам"
        
        
        ws1.append(["Год", "Средняя зарплата","Средняя зарплата - " + self.name, "Количество вакансий", "Количество вакансий - " + self.name])
        self.do_bold(ws1,"A1:E1")
        self.do_border(ws1,"A1:E"+str(len(self.salarysYear.keys())+1))
        count = 2
        for year in self.salarysYear.keys():
            ws1['A' + str(count)] = year
            ws1['B' + str(count)] = self.salarysYear[year]
            ws1['C' + str(count)] = self.filterSalarysYear[year]
            ws1['D' + str(count)] = self.countVacancyesYear[year]
            ws1['E' + str(count)] = self.filterCountVacancyesYear[year]

            count += 1
        
    
        ws2 = wb.create_sheet("Статистика по городам")           
        
        
        
        ws2.append(["Город", "Уровень зараплат" , "", "Город", "Доля вакансий"])
        self.do_bold(ws2,"A1:E1")
        self.do_border(ws2,"A1:B"+str(len(self.salaryTown.keys())+1))
        
        count = 2
        for town in self.salaryTown.keys():
            ws2['A' + str(count)] = town
            ws2['B' + str(count)] = self.salaryTown[town]
            count += 1
        self.do_border(ws2,"D1:E"+str(len(self.upgradeVacanciesTown.keys())+1))
        count = 2    
        for town in self.upgradeVacanciesTown.keys():
            ws2['D' + str(count)] = town
            ws2['E' + str(count)] = str(round(float(self.upgradeVacanciesTown[town])*100,2))+"%"

            count += 1
        self.do_width_sizes(ws1)
        self.do_width_sizes(ws2)
        wb.save('report.xlsx')
    def createPdf(self):
        """Создаёт документ pdf
        """
        name = self.name

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("index.html")

        pdf_template = template.render({'name': name})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        options = {'enable-local-file-access': None}
        pdf_template = pdf_template.replace("$path", os.path.abspath(os.curdir)+"\\" )
        
        htmlTable = "<table class='main_table'><tr><th>Год</th><th>Средняя зарплата</th><th>Средняя зарплата - "
        htmlTable += self.name + "</th><th>Количество вакансий</th><th>Количество вакансий - " + self.name + "</th></tr>"
        for index in range(len(list(self.salarysYear.keys()))):
            htmlTable += "<tr>"
            htmlTable += ("<td>" + str(list(self.salarysYear.keys())[index]) + "</td>")
            htmlTable += ("<td>" + str(list(self.salarysYear.values())[index]) + "</td>")
            htmlTable += ("<td>" + str(list(self.countVacancyesYear.values())[index]) + "</td>")
            htmlTable += ("<td>" + str(list(self.filterSalarysYear.values())[index]) + "</td>")
            htmlTable += ("<td>" + str(list(self.filterCountVacancyesYear.values())[index]) + "</td>")
            htmlTable += "</tr>"
        htmlTable += '</tr></table><h1 align="center" >Статистика по городам</h1>'
        
        htmlTable += "<table class='tableTownSalarys'><tr><th>Город</th><th>Уровень зарплат</th>"
        for index in range(len(list(self.salaryTown.keys()))):
            htmlTable += "<tr>"
            htmlTable += ("<td>" + str(list(self.salaryTown.keys())[index]) + "</td>")
            htmlTable += ("<td>" + str(list(self.salaryTown.values())[index]) + "</td>")
            htmlTable += "</tr>"
        htmlTable += "</tr></table>"
        
        htmlTable += "<table class='tableTownProcent'><tr><th>Город</th><th>Уровень зарплат</th>"
        for index in range(len(list(self.upgradeVacanciesTown.keys()))):
            htmlTable += "<tr>"
            htmlTable += ("<td>" + str(list(self.upgradeVacanciesTown.keys())[index]) + "</td>")
            htmlTable += ("<td>" + str(round(list(self.upgradeVacanciesTown.values())[index] * 100, 2)) + "%" + "</td>")
            htmlTable += "</tr>"
        htmlTable += "</tr></table>"

        pdf_template = pdf_template.replace("$tables", htmlTable)
        pdfkit.from_string(pdf_template, 'report.pdf', options=options, configuration=config)
if __name__=="__main__":

    doctest.testmod()
    whatPrint=input("Выбери что вывести 'Вакансии' или 'Статистика': ")
    ourInput = InputConect(whatPrint)
    """Пошаговый алгоритм условий и запуска функций при разных требованиях к выводу
    """
    if whatPrint=="Статистика":
        ourInput.checkInput()
        vacancies=DataSet(ourInput.file,[""], "","")

        vacancies.correctVacanceis()
        salarysYear,countVacancyesYear,VacanciesTown,salaryTown,filterSalarysYear,filterCountVacancyesYear=vacancies.yearDinamic(ourInput.filterElements[1])

        salarysYearKey=list(salarysYear.keys())
        VacanciesTownKey=list(VacanciesTown.keys())
        salaryTownKey=list(salaryTown.keys())

        upgradeVacanciesTown={}

        filterSalarysYearKey=list(filterSalarysYear.keys())

        filterSalaryTownKey=list(filterSalarysYear.keys())
        i=0
        while(i<10):
            if i<len(salarysYearKey):
                salarysYear[salarysYearKey[i]]=int(salarysYear[salarysYearKey[i]]/countVacancyesYear[salarysYearKey[i]])
            if i<len(salaryTownKey):
                salaryTown[salaryTownKey[i]]=int(salaryTown[salaryTownKey[i]]/VacanciesTown[salaryTownKey[i]])
            if i<len(filterSalarysYearKey) and filterCountVacancyesYear[filterSalarysYearKey[i]]!=0:
                filterSalarysYear[filterSalarysYearKey[i]]=int(filterSalarysYear[filterSalarysYearKey[i]]/filterCountVacancyesYear[filterSalarysYearKey[i]])
            if i<len(VacanciesTownKey):
                proc=round(VacanciesTown[VacanciesTownKey[i]]/len(vacancies.vacancies_objects),4)
                if proc>=0.01:
                    upgradeVacanciesTown[VacanciesTownKey[i]]=proc
            i+=1
        salarysYear=dict(list(salarysYear.items())[0:10])
        countVacancyesYear=   dict(list(countVacancyesYear.items())[0:10])
        filterSalarysYear=dict(list(filterSalarysYear.items())[0:10])
        filterCountVacancyesYear=dict(list(filterCountVacancyesYear.items())[0:10])
        salaryTown=dict(list(salaryTown.items())[0:10])
        upgradeVacanciesTown=dict(list(upgradeVacanciesTown.items())[0:10])
        print("Динамика уровня зарплат по годам: ",end="")
        print(salarysYear)
        print("Динамика количества вакансий по годам: ",end="")
        print(countVacancyesYear)
        print("Динамика уровня зарплат по годам для выбранной профессии: ",end="")
        print(filterSalarysYear)
        print("Динамика количества вакансий по годам для выбранной профессии: ",end="")
        print(filterCountVacancyesYear)
        print("Уровень зарплат по городам (в порядке убывания): ",end="")
        print(salaryTown)
        print("Доля вакансий по городам (в порядке убывания): ",end="")
        print(upgradeVacanciesTown)
        exel = Report(ourInput.filterElements[1],salarysYear,countVacancyesYear,filterSalarysYear,filterCountVacancyesYear,salaryTown,upgradeVacanciesTown)
        exel.generate_excel()
        exel.generate_diagrams()
        exel.createPdf()
    if whatPrint=="Вакансии":
        ourInput.checkInput()
        vacancies=DataSet(ourInput.file,ourInput.filterElements, ourInput.sortElements,ourInput.reversVacancies)
        ourInput.print_vacancies(vacancies.correctVacanceis(),fieldToRus)